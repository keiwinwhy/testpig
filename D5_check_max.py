from openpyxl import load_workbook
from pywinauto import *
from time import sleep
import xlrd
from datetime import datetime
import openpyxl
import os

from pywinauto import keyboard

from getdata_sf import get_value
from D5_printlabel import action_printlabel

parent_dir = os.path.abspath(os.pardir)
current_dir = os.getcwd()
current_date = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
print(current_date)

t = 1
t1 = 3
t2 = 2

app = Application(backend="uia").start('control.exe')
cpanel = Desktop(backend="uia").ControlPanel

sleep(t1)

cpanel.child_window(title="Hardware and Sound").click_input()
app = Application(backend="uia").connect(title_re=".*Hardware and Sound*")
dp = app.window(best_match='Hardware and Sound', top_level_only=False)
dp.child_window(title="Devices and Printers").click_input()

sleep(t1)

app2 = Application(backend="uia").connect(title_re=".*Devices and Printers*")
dp2 = app2.window(best_match="Devices and Printers",top_level_only=False)
#dp2.print_control_identifiers()

result_name = current_dir + "\\" + 'Restul_SRSmapping_Forerunner&Stingray.xlsx'
# 写入指定文件
wb = load_workbook(filename=result_name)  ##读取路径
ws = wb.active

file_name = current_dir + "\\" + 'SRSmapping_Forerunner&Stingray.xlsx'
book = xlrd.open_workbook(file_name)
work_sheet_range = range(book.nsheets)
try:
    work_sheet = book.sheet_by_name("Stringray mapping V5")
except:
    print("work sheet not found is %s name" %file_name)

num_cols = work_sheet.ncols
try:
    #for j in range(2, num_cols):
    for j in range(2, 5):
        prt_type = work_sheet.cell_value(1,j)
        prtname = work_sheet.cell_value(4,j)
        pre_max_width = str(int(work_sheet.cell_value(9,j)))
        pre_max_height = str(int(work_sheet.cell_value(10,j)))

        print("prt_type: ", prt_type)
        print("prtname: ", prtname)
        print("pre_max_width: ", pre_max_width)
        print("pre_max_height: ", pre_max_height)
        print("\n")

        dp2.child_window(title = prtname).click_input(button = 'right')
        app2.Context.Printingpreferences.click_input()
        sleep(t1)

        app3 = Application(backend = "uia").connect(title_re = prtname + "*")
        dp3 = app3.window(best_match = prtname + " Printing Preferences",top_level_only = False)
    #    dp3.print_control_identifiers()

        dp3.child_window(title="mm", control_type="RadioButton").click_input()
        dp3 = app3.window(best_match = prtname + " Printing Preferences",top_level_only = False)
    #    dp3.print_control_identifiers()


        dp3.child_window(title = "Width:", control_type = "Edit").click_input()
        for i in range(0,7):
            keyboard.send_keys("{BACKSPACE}")
            sleep(t)
            
        dp3.child_window(title = "Width:", control_type = "Edit").type_keys("80000")
        sleep(t)
        
        dp3.child_window(title = "Height:", control_type = "Edit").click_input()
        for i in range(0,7):
            keyboard.send_keys("{BACKSPACE}")
            sleep(t)

        dp3.child_window(title = "Height:", control_type = "Edit").type_keys("80000")    
        sleep(t)

        dp3.child_window(title = "Apply", control_type = "Button").click()
        sleep(t)

        #截图
        a = dp3.capture_as_image()
        a.save( current_dir + "\\" + prtname + "_max_" + current_date + ".png")
        sleep(t)

        app3 = Application(backend = "uia").connect(title_re = prtname + "*")
        dp3 = app3.window(best_match = prtname + " Printing Preferences",top_level_only = False)

        max_width = dp3.child_window(title = "Width:", control_type = "Edit").get_value()
        max_height = dp3.child_window(title = "Height:", control_type = "Edit").get_value()
        t_max_width = str(int(float(max_width) * 10))
        t_max_height = str(int(float(max_height) * 10))

        dp3.child_window(title="Cancel", control_type="Button").click_input()
        sleep(t1)

        app2 = Application(backend="uia").connect(title_re=".*Devices and Printers*")
        dp2 = app2.window(best_match="Devices and Printers", top_level_only=False)
        sleep(t1)

        # print test label
        action_printlabel(app2, dp2, prtname)

        #结果判断
        v_max_width = ''
        v_max_height = ''

        if pre_max_width == t_max_width:
            v_max_width = "E[" + pre_max_width + "]"
        else:
            v_max_width = "NE[expected value:" + pre_max_width + ", actual value:" + t_max_width + "]"

        if pre_max_height == t_max_height:
            v_max_height = "E[" + pre_max_height + "]"
        else:
            v_max_height = "NE[expected value:" + pre_max_height + ", actual value:" + t_max_height + "]"

        # read the save file content
        p_width, p_dark, p_speed, p_height = get_value(prt_type)
        v_max_width = v_max_width + "\n" + p_width
        v_max_height = v_max_height + "\n" + p_height

        #结果值写入表格
        # 指定位置为指定列中的不同的行，值为value(行列值要各自加1,从1开始，非0)
        ws.cell(row=10, column=j+1, value=v_max_width)
        ws.cell(row=11, column=j+1, value=v_max_height)


except Exception as e:
    print(e)

wb.save(result_name)
dp.type_keys("%{F4}")
