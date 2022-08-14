from openpyxl import load_workbook
from pywinauto import *
from time import sleep
import xlrd
from datetime import datetime
import openpyxl
import os

parent_dir = os.path.abspath(os.pardir)
current_dir = os.getcwd()
current_date = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
print(current_date)

t = 1
t1 = 3
t2 = 2

app = Application(backend="uia").start('control.exe')
cpanel = Desktop(backend="uia").ControlPanel

sleep(t1)

cpanel.child_window(title="Hardware and Sound").click_input()
app = Application(backend="uia").connect(title_re=".*Hardware and Sound*")
dp = app.window(best_match='Hardware and Sound', top_level_only=False)
dp.child_window(title="Devices and Printers").click_input()

sleep(t1)

app2 = Application(backend="uia").connect(title_re=".*Devices and Printers*")
dp2 = app2.window(best_match="Devices and Printers",top_level_only=False)
#dp2.print_control_identifiers()

result_name = current_dir + "\\" + 'Restul_SRSmapping_Forerunner&Stingray.xlsx'
# 写入指定文件
wb = load_workbook(filename=result_name)  ##读取路径
ws = wb.active

file_name = current_dir + "\\" + 'SRSmapping_Forerunner&Stingray.xlsx'
book = xlrd.open_workbook(file_name)
work_sheet_range = range(book.nsheets)
try:
    work_sheet = book.sheet_by_name("Stringray mapping V5")
except:
    print("work sheet not found is %s name" %file_name)

num_cols = work_sheet.ncols
#for j in range(2, num_cols):
for j in range(2, 7):
    prt_type = work_sheet.cell_value(1,j)
    prtname = work_sheet.cell_value(4,j)
    pre_default_width = str(int(work_sheet.cell_value(5,j)))
    pre_default_height = str(int(work_sheet.cell_value(6,j)))
    pre_default_speed = str(int(work_sheet.cell_value(7,j)))
    pre_default_darkness = str(int(work_sheet.cell_value(8,j)))
    pre_SpeedLimits = str(work_sheet.cell_value(13,j))
    pre_DarknessLimits = str(work_sheet.cell_value(14,j))

    print("prt_type: ", prt_type)
    print("prtname: ", prtname)
    print("pre_default_width: ", pre_default_width)
    print("pre_default_height: ", pre_default_height)
    print("pre_default_speed: ", pre_default_speed)
    print("pre_default_darkness: ", pre_default_darkness)
    print("pre_SpeedLimits: ", pre_SpeedLimits)
    print("pre_DarknessLimits: ", pre_DarknessLimits)
    print("\n")

    dp2.child_window(title = prtname).click_input(button = 'right')
    app2.Context.Printingpreferences.click_input()
    sleep(t1)

    app3 = Application(backend = "uia").connect(title_re = prtname + "*")
    dp3 = app3.window(best_match = prtname + " Printing Preferences",top_level_only = False)
#    dp3.print_control_identifiers()

    dp3.child_window(title="mm", control_type="RadioButton").click_input()
    dp3 = app3.window(best_match = prtname + " Printing Preferences",top_level_only = False)
#    dp3.print_control_identifiers()

    #截图
    a = dp3.capture_as_image()
    print (2)
    a.save( current_dir + "\\" + prtname + "_default_" + current_date + ".png")

    text_width = dp3.child_window(title = "Width:", control_type = "Edit").get_value()
    t_width = str(int(float(text_width) * 10))
    print("t_width：", t_width)

    text_height = dp3.child_window(title = "Height:", control_type = "Edit").get_value()
    t_height = str(int(float(text_height) * 10))
    print("t_height：", t_height)

    dp3.child_window(title="inch", control_type="RadioButton").click_input()
    dp3 = app3.window(best_match = prtname + " Printing Preferences",top_level_only = False)
#    dp3.print_control_identifiers()

    # get speed default value
    t_speed = dp3.child_window(title="Speed:", control_type="ComboBox").selected_text()
    print("speed default value: ", t_speed)
    print(type(t_speed), "\n")

    dp3.child_window(title="Speed:", control_type="ComboBox").click_input()
    dp3 = app3.window(best_match=prtname + " Printing Preferences", top_level_only=False)
    #dp3.print_control_identifiers()

    i4 = dp3.child_window(title="Speed:", control_type="List").item_count()
    print("SpeedLimits length: ", i4)

    i5 = dp3.child_window(title="Speed:", control_type="List").texts()
    t_SpeedLimits = ''
    for i in range(0,i4):
        #print("i5[i]: ", i5[i][0])
        #print(type(i5[i][0]))
        if i != i4-1:
            t_SpeedLimits = t_SpeedLimits + str(i5[i][0]) + ","
        else:
            t_SpeedLimits = t_SpeedLimits + str(i5[i][0])

    print("t_SpeedLimits: ", t_SpeedLimits)
    #print(type(SpeedLimits),"\n")

    dp3.child_window(title="Speed:", control_type="ComboBox").click_input()
    dp3 = app3.window(best_match=prtname + " Printing Preferences", top_level_only=False)

    # get darkness default value
    t_darkness = dp3.child_window(title="Darkness:", control_type="ComboBox").selected_text()
    print("darkness_default: ", t_darkness)
    print(type(t_darkness), "\n")

    dp3.child_window(title="Darkness:", control_type="ComboBox").click_input()
    dp3 = app3.window(best_match=prtname + " Printing Preferences", top_level_only=False)

    j4 = dp3.child_window(title="Darkness:", control_type="List").texts()
    #print("DarknessList: ", j4)
    #print(type(j4),"\n")

    t_DarknessLimits = str(j4[0][0]) + "-" + str(j4[len(j4)-1][0])
    print("t_DarknessLimits: ", t_DarknessLimits)

    dp3.child_window(title="Darkness:", control_type="ComboBox").click_input()
    dp3 = app3.window(best_match=prtname + " Printing Preferences", top_level_only=False)
    dp3.child_window(title="Cancel", control_type="Button").click_input()
    sleep(t1)

    app2 = Application(backend="uia").connect(title_re=".*Devices and Printers*")
    dp2 = app2.window(best_match="Devices and Printers", top_level_only=False)
    sleep(t1)

    #结果判断
    v_width = ''
    v_height = ''
    v_speed = ''
    v_darkness = ''
    v_SpeedLimits = ''
    v_DarknessLimits = ''

    if pre_default_width == t_width:
        v_width = "P[" + pre_default_width + "]"
    else:
        v_width = "F[expected value:" + pre_default_width + ", actual value:" + t_width + "]"

    if pre_default_height == t_height:
        v_height = "P[" + pre_default_height + "]"
    else:
        v_height = "F[expected value:" + pre_default_height + ", actual value:" + t_height + "]"

    if pre_default_speed == t_speed:
        v_speed = "P[" + pre_default_speed + "]"
    else:
        v_speed = "F[expected value:" + pre_default_speed + ", actual value:" + t_speed + "]"

    if pre_default_darkness == t_darkness:
        v_darkness = "P[" + pre_default_darkness + "]"
    else:
        v_darkness = "F[expected value:" + pre_default_darkness + ", actual value:" + t_darkness + "]"

    if pre_SpeedLimits == t_SpeedLimits:
        v_SpeedLimits = "P[" + pre_SpeedLimits + "]"
    else:
        v_SpeedLimits = "F[expected value:" + pre_SpeedLimits + ", actual value:" + t_SpeedLimits + "]"

    if pre_DarknessLimits == t_DarknessLimits:
        v_DarknessLimits = "P[" + pre_DarknessLimits + "]"
    else:
        v_DarknessLimits = "F[expected value:" + pre_DarknessLimits + ", actual value:" + t_DarknessLimits + "]"

    #结果值写入表格
    # 指定位置为指定列中的不同的行，值为value(行列值要各自加1,从1开始，非0)
    ws.cell(row=6, column=j+1, value=v_width)
    ws.cell(row=7, column=j+1, value=v_height)
    ws.cell(row=8, column=j+1, value=v_speed)
    ws.cell(row=9, column=j+1, value=v_darkness)
    ws.cell(row=14, column=j+1, value=v_SpeedLimits)
    ws.cell(row=15, column=j+1, value=v_DarknessLimits)

wb.save(result_name)
dp.type_keys("%{F4}")
